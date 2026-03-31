"""
Módulo para exportar um único arquivo Excel com duas abas:
- ERROS   (ID_RELATORIO, EQUIPE, DATA, ERRO, CATEGORIA, ORIGEM)
- STATUS  (ID_RELATORIO, EQUIPE, DATA, STATUS, ORIGEM)
Preserva dados manuais e GARANTE a coluna STATUS.
"""

import sqlite3
import pandas as pd
from datetime import datetime
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_NAME = "banco.db"
EXCEL_UNIFICADO = "relatorios_completos.xlsx"
BACKUP_EXCEL = "relatorios_completos_backup.xlsx"

# ----------------------------------------------------------------------
# 1. Obter dados do sistema (banco de dados)
# ----------------------------------------------------------------------
def obter_dados_erros_sistema():
    try:
        conn = sqlite3.connect(DB_NAME)
        query = """
        SELECT 
            r.id as ID_RELATORIO,
            r.equipe as EQUIPE,
            r.data as DATA,
            r.status as STATUS,
            r.data_criacao as DATA_CRIACAO,
            e.descricao as ERRO,
            ep.categoria as CATEGORIA,
            1 as QUANTIDADE
        FROM relatorios r
        LEFT JOIN erros e ON r.id = e.relatorio_id
        LEFT JOIN erros_padrao ep ON e.descricao = ep.descricao
        WHERE e.descricao IS NOT NULL
        ORDER BY r.data_criacao DESC, r.id DESC
        """
        df = pd.read_sql(query, conn)
        conn.close()
        if df.empty:
            return pd.DataFrame(columns=['ID_RELATORIO','EQUIPE','DATA','STATUS','DATA_CRIACAO','ERRO','CATEGORIA','QUANTIDADE'])
        df['DATA'] = pd.to_datetime(df['DATA'], format='%d/%m/%Y', errors='coerce').dt.strftime('%d/%m/%Y')
        df['DATA_CRIACAO'] = pd.to_datetime(df['DATA_CRIACAO']).dt.strftime('%d/%m/%Y %H:%M')
        df['ORIGEM'] = 'SISTEMA'
        return df
    except Exception as e:
        print(f"Erro erros sistema: {e}")
        return pd.DataFrame()

def obter_dados_status_sistema():
    try:
        conn = sqlite3.connect(DB_NAME)
        query = """
        SELECT 
            r.id as ID_RELATORIO,
            r.equipe as EQUIPE,
            r.data as DATA,
            r.status as STATUS,
            r.data_criacao as DATA_CRIACAO,
            COUNT(e.id) as TOTAL_ERROS,
            GROUP_CONCAT(e.descricao, '; ') as LISTA_ERROS
        FROM relatorios r
        LEFT JOIN erros e ON r.id = e.relatorio_id
        GROUP BY r.id, r.equipe, r.data, r.status, r.data_criacao
        ORDER BY r.data_criacao DESC, r.id DESC
        """
        df = pd.read_sql(query, conn)
        conn.close()
        if df.empty:
            return pd.DataFrame(columns=['ID_RELATORIO','EQUIPE','DATA','STATUS','DATA_CRIACAO','TOTAL_ERROS','LISTA_ERROS'])
        df['DATA'] = pd.to_datetime(df['DATA'], format='%d/%m/%Y', errors='coerce').dt.strftime('%d/%m/%Y')
        df['DATA_CRIACAO'] = pd.to_datetime(df['DATA_CRIACAO']).dt.strftime('%d/%m/%Y %H:%M')
        df['ORIGEM'] = 'SISTEMA'
        return df
    except Exception as e:
        print(f"Erro status sistema: {e}")
        return pd.DataFrame()

# ----------------------------------------------------------------------
# 2. Ler dados manuais (com segurança)
# ----------------------------------------------------------------------
def ler_dados_manuais_seguro(arquivo, nome_aba):
    if not os.path.exists(arquivo):
        return pd.DataFrame()
    try:
        df = pd.read_excel(arquivo, sheet_name=nome_aba, dtype=str)
        if df.empty:
            return pd.DataFrame()
        if 'ORIGEM' not in df.columns:
            df['ORIGEM'] = 'MANUAL'
        return df[df['ORIGEM'] == 'MANUAL'].copy()
    except Exception as e:
        print(f"Erro ao ler {nome_aba}: {e}")
        return pd.DataFrame()

# ----------------------------------------------------------------------
# 3. Exportação principal
# ----------------------------------------------------------------------
def exportar_planilha_completa():
    print("="*60)
    print("📊 GERANDO EXCEL COM COLUNA STATUS GARANTIDA")
    print("="*60)

    # Backup
    if os.path.exists(EXCEL_UNIFICADO):
        try:
            shutil.copy2(EXCEL_UNIFICADO, BACKUP_EXCEL)
            print(f"✅ Backup criado: {BACKUP_EXCEL}")
        except Exception as e:
            print(f"⚠️ Backup falhou: {e}")

    # ---------- ERROS ----------
    df_erros_sis = obter_dados_erros_sistema()
    df_erros_man = ler_dados_manuais_seguro(EXCEL_UNIFICADO, "ERROS")

    if df_erros_man.empty and df_erros_sis.empty:
        df_erros = pd.DataFrame(columns=['ID_RELATORIO','EQUIPE','DATA','ERRO','CATEGORIA','ORIGEM'])
    elif not df_erros_man.empty and df_erros_sis.empty:
        df_erros = df_erros_man
    elif df_erros_man.empty and not df_erros_sis.empty:
        df_erros = df_erros_sis[['ID_RELATORIO','EQUIPE','DATA','ERRO','CATEGORIA','ORIGEM']]
    else:
        # Mescla
        df_erros_man['CHAVE'] = df_erros_man['ID_RELATORIO'].astype(str)+'_'+df_erros_man['ERRO'].astype(str)
        df_erros_sis['CHAVE'] = df_erros_sis['ID_RELATORIO'].astype(str)+'_'+df_erros_sis['ERRO'].astype(str)
        df_erros_sis_novos = df_erros_sis[~df_erros_sis['CHAVE'].isin(df_erros_man['CHAVE'])]
        df_erros_sis_novos.drop('CHAVE', axis=1, inplace=True)
        df_erros_man.drop('CHAVE', axis=1, inplace=True)
        df_erros = pd.concat([df_erros_man, df_erros_sis_novos], ignore_index=True)

    # Garantir colunas de ERROS
    col_erros = ['ID_RELATORIO','EQUIPE','DATA','ERRO','CATEGORIA','ORIGEM']
    for c in col_erros:
        if c not in df_erros.columns:
            df_erros[c] = None
    df_erros = df_erros[col_erros]

    # ---------- STATUS ----------
    df_status_sis = obter_dados_status_sistema()
    df_status_man = ler_dados_manuais_seguro(EXCEL_UNIFICADO, "STATUS")

    if df_status_man.empty and df_status_sis.empty:
        df_status = pd.DataFrame(columns=['ID_RELATORIO','EQUIPE','DATA','STATUS','ORIGEM'])
    elif not df_status_man.empty and df_status_sis.empty:
        df_status = df_status_man
    elif df_status_man.empty and not df_status_sis.empty:
        df_status = df_status_sis[['ID_RELATORIO','EQUIPE','DATA','STATUS','ORIGEM']]
    else:
        ids_man = df_status_man['ID_RELATORIO'].astype(str).tolist()
        df_status_sis_novos = df_status_sis[~df_status_sis['ID_RELATORIO'].astype(str).isin(ids_man)]
        df_status = pd.concat([df_status_man, df_status_sis_novos], ignore_index=True)

    # FORÇAR A EXISTÊNCIA DA COLUNA STATUS
    col_status = ['ID_RELATORIO','EQUIPE','DATA','STATUS','ORIGEM']
    for c in col_status:
        if c not in df_status.columns:
            df_status[c] = None
    # Se a coluna STATUS existe mas está vazia, preencher com base no STATUS do sistema? 
    # Não, manter o que vier. Mas garantir que a coluna está lá.
    df_status = df_status[col_status]

    # DEBUG: mostrar colunas no console
    print(f"Colunas da aba STATUS: {list(df_status.columns)}")
    if 'STATUS' not in df_status.columns:
        print("⚠️ ERRO: Coluna STATUS ainda não foi criada! Adicionando forçadamente...")
        df_status['STATUS'] = None

    # ---------- ESCREVER EXCEL ----------
    try:
        with pd.ExcelWriter(EXCEL_UNIFICADO, engine='openpyxl') as writer:
            df_erros.to_excel(writer, sheet_name='ERROS', index=False)
            df_status.to_excel(writer, sheet_name='STATUS', index=False)

            # Formatação
            for sheet_name in ['ERROS','STATUS']:
                ws = writer.sheets[sheet_name]
                # Cabeçalho
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                # Bordas
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border
                ws.auto_filter.ref = ws.dimensions
                # Ajuste de largura
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_len:
                                max_len = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = min(max_len+2, 50)

        print("✅ Excel gerado com sucesso!")
        print(f"   Aba STATUS contém as colunas: {', '.join(df_status.columns)}")
        return {'sucesso': True, 'arquivo': EXCEL_UNIFICADO, 'caminho': os.path.abspath(EXCEL_UNIFICADO)}
    except Exception as e:
        print(f"❌ Erro: {e}")
        return {'sucesso': False, 'erro': str(e)}

def exportar_dados_para_excel_fixo():
    return exportar_planilha_completa()

if __name__ == "__main__":
    exportar_dados_para_excel_fixo()